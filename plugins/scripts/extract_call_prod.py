import os

import pandas as pd
from airflow.sdk import task
from openpyxl import load_workbook
from utils.logger import get_logger

logger = get_logger(__name__)


def _define_tables(path: str):
    """
    Dynamic filters for tables between none lines
    Supposed each data have the sames titles
    :return dict []
        .keys()
    """
    if not os.path.exists(path):
        logger.error(f"File not found in {path}")
        raise FileNotFoundError(f"File not found in {path}")
    try:
        logger.info(f"Extracting data from {path}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        ws = workbook.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        logger.error(f"Error extracting data from {path}: {str(e)}")
        raise e

    final_tables = {}
    actual_bloc = []

    for row in rows:
        is_all_none = all(value is None for value in row)
        if is_all_none:
            if actual_bloc:
                title = str(actual_bloc[0][0]).strip() if actual_bloc[0][0] else None
                final_tables[title] = actual_bloc
                actual_bloc = []
            continue
        actual_bloc.append(row)
    if actual_bloc:
        title = str(actual_bloc[0][0]).strip() if actual_bloc[0][0] else "last table"
        final_tables[title] = actual_bloc
    return final_tables


@task
def extract_call_prod(BASE_DIR: str, ds: str) -> str:
    """
    Extract call production by using dynamic filters for tables between none lines
    :param BASE_DIR:
    :param ds: Jinja template
    :return: str file path of the processed data
     :raises : FileNotFoundError: if the raw excel file is not found in the specified path
     :raises : Exception: if there is any error during the extraction process
    """

    PROCESSED_TABLES = [
        "Traités",
        "Durée moy de traitement",
        "Durée moy com",
        "Sortants effectués",
    ]
    month, year = "01", "2023"
    dataframes = {}
    data = _define_tables(
        os.path.join(BASE_DIR, "raw", "STAT APPEL GLOBAL FR BEFR.xlsx")
    )
    selected_tables = {}
    for table_title in data.keys():
        for table in PROCESSED_TABLES:
            if table in table_title:
                selected_tables[table] = data[table_title]
    for title, lines in selected_tables.items():
        if len(lines) > 1:
            df = pd.DataFrame(lines[1:], columns=lines[0])
            df = df.rename(columns={lines[0][0]: "Agent"})
            df["Agent"] = (
                df["Agent"].astype(str).str.lower().str.strip().str.split(" ").str[0]
            )
            df = df[df["Agent"].str.startswith("adt_", na=False)]
            date_mapping = {
                col: pd.to_datetime(
                    f"{int(col):02d}/{month}/{year}", format="%d/%m/%Y", errors="coerce"
                )
                for col in df.columns
                if str(col).isdigit()
            }
            df = df.rename(columns=date_mapping)
            date_cols = [col for col in df.columns if isinstance(col, pd.Timestamp)]
            """format data from wide to long for each table"""
            df = df.melt(
                id_vars=["Agent"],
                value_vars=date_cols,
                var_name="Date",
                value_name=title,
            )
            dataframes[title] = df

    df_traite = dataframes["Traités"]
    df_traitement = dataframes["Durée moy de traitement"]
    df_comm = dataframes["Durée moy com"]
    df_sortant = dataframes["Sortants effectués"]

    # ------------------- TRAITEE
    column_mapping = {
        "Traités": "Appels traités",
    }
    df_traite = df_traite.rename(columns=column_mapping)

    # ------------------- DUREE MOY DE TRAITEMENT
    df_traitement["Durée moy de traitement"] = pd.to_timedelta(
        df_traitement["Durée moy de traitement"]
    ).dt.total_seconds()

    # ------------------- DUREE MOY DE COM
    df_comm["Durée moy com"] = pd.to_timedelta(
        df_comm["Durée moy com"]
    ).dt.total_seconds()

    # ------------------- SORTANTS EFFECTUES
    column_mapping = {"Sortants effectués": "Appels sortants"}
    df_sortant = df_sortant.rename(columns=column_mapping)

    df_call_prod = (
        df_traite.merge(df_traitement, on=["Agent", "Date"], how="left")
        .merge(df_comm, on=["Agent", "Date"], how="left")
        .merge(df_sortant, on=["Agent", "Date"], how="left")
    )

    processed_dir = os.path.join(BASE_DIR, "tmp")
    os.makedirs(processed_dir, exist_ok=True)
    output_filename = f"processed_call_prod_{ds}.parquet"
    output_path = os.path.join(processed_dir, output_filename)

    df_call_prod.to_parquet(output_path, index=False)

    logger.info(
        f"Processed data saved to {output_path} - {df_call_prod.shape[0]} rows and - {df_call_prod.shape[1]} columns"
    )
    return output_path
